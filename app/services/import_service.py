"""Service for importing animals from CSV and Excel files."""
from __future__ import annotations

import csv
import os
import shutil
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import app_config
from services.animal_service import AnimalService
from storage.file_store import get_file_store


@dataclass
class ImportError:
    """Represents an error during import."""
    row: int
    message: str


@dataclass
class ImportResult:
    """Result of an import operation."""
    success_count: int = 0
    errors: List[ImportError] = field(default_factory=list)
    
    @property
    def total_rows(self) -> int:
        return self.success_count + len(self.errors)
    
    @property
    def has_errors(self) -> bool:
        return len(self.errors) > 0
    
    @property
    def all_failed(self) -> bool:
        return self.success_count == 0 and len(self.errors) > 0


class ImportService:
    """Service for importing animals from CSV and Excel files."""
    
    # Valid values for validation
    VALID_ANIMAL_TYPES = {"dog", "cat", "other"}
    VALID_HEALTH_STATUSES = {"healthy", "recovering", "injured"}
    MIN_AGE = 0
    MAX_AGE = 21
    
    def __init__(self, db_path: Optional[str] = None) -> None:
        self.animal_service = AnimalService(db_path or app_config.DB_PATH)
        self.file_store = get_file_store()
    
    def import_from_file(self, file_path: str) -> ImportResult:
        """Import animals from a CSV or Excel file.
        
        Args:
            file_path: Path to the CSV or Excel file
            
        Returns:
            ImportResult with success count and list of errors
        """
        ext = Path(file_path).suffix.lower()
        
        if ext == ".csv":
            return self._import_from_csv(file_path)
        elif ext in (".xlsx", ".xls"):
            return self._import_from_excel(file_path)
        else:
            result = ImportResult()
            result.errors.append(ImportError(
                row=0,
                message=f"Unsupported file format '{ext}'. Use .csv or .xlsx"
            ))
            return result
    
    def _import_from_csv(self, file_path: str) -> ImportResult:
        """Import animals from a CSV file."""
        result = ImportResult()
        file_dir = os.path.dirname(os.path.abspath(file_path))
        
        try:
            with open(file_path, "r", encoding="utf-8-sig") as f:
                lines = [line for line in f if not line.strip().startswith("#")]
                
            if not lines:
                result.errors.append(ImportError(row=0, message="File is empty or contains only comments"))
                return result
            
            # Parse CSV from filtered lines
            reader = csv.DictReader(lines)
            
            headers = reader.fieldnames or []
            header_error = self._validate_headers(headers)
            if header_error:
                result.errors.append(ImportError(row=1, message=header_error))
                return result
            
            for row_num, row in enumerate(reader, start=2):  # Start at 2 (1 is header)
                self._process_row(row_num, row, file_dir, result)
                
        except FileNotFoundError:
            result.errors.append(ImportError(row=0, message=f"File not found: {file_path}"))
        except csv.Error as e:
            result.errors.append(ImportError(row=0, message=f"CSV parsing error: {e}"))
        except Exception as e:
            result.errors.append(ImportError(row=0, message=f"Unexpected error: {e}"))
        
        return result
    
    def _import_from_excel(self, file_path: str) -> ImportResult:
        """Import animals from an Excel file."""
        result = ImportResult()
        file_dir = os.path.dirname(os.path.abspath(file_path))
        
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            result.errors.append(ImportError(
                row=0,
                message="Excel support requires openpyxl. Install with: pip install openpyxl"
            ))
            return result
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            if "Animals" in wb.sheetnames:
                ws = wb["Animals"]
            else:
                ws = wb.active
            
            if ws is None:
                result.errors.append(ImportError(row=0, message="No worksheet found in Excel file"))
                return result
            
            rows = list(ws.iter_rows(values_only=True))
            
            if not rows:
                result.errors.append(ImportError(row=0, message="Excel file is empty"))
                return result
            
            # First row is headers
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            header_error = self._validate_headers(headers)
            if header_error:
                result.errors.append(ImportError(row=1, message=header_error))
                return result
            
            col_map = {name: idx for idx, name in enumerate(headers)}
            
            for row_num, row_data in enumerate(rows[1:], start=2):
                # Convert row to dict
                row = {}
                for col_name in ["name", "animal_type", "breed", "age", "health_status", "photo"]:
                    if col_name in col_map:
                        value = row_data[col_map[col_name]]
                        row[col_name] = str(value).strip() if value is not None else ""
                    else:
                        row[col_name] = ""
                
                # Skip empty rows
                if not any(row.values()):
                    continue
                
                self._process_row(row_num, row, file_dir, result)
            
            wb.close()
            
        except FileNotFoundError:
            result.errors.append(ImportError(row=0, message=f"File not found: {file_path}"))
        except Exception as e:
            result.errors.append(ImportError(row=0, message=f"Excel parsing error: {e}"))
        
        return result
    
    def _validate_headers(self, headers: List[str]) -> Optional[str]:
        """Validate that required headers are present. Returns error message or None."""
        headers_lower = [h.lower().strip() for h in headers]
        
        required = {"name", "animal_type", "age", "health_status"}
        missing = required - set(headers_lower)
        
        if missing:
            return f"Missing required columns: {', '.join(sorted(missing))}"
        
        return None
    
    def _process_row(self, row_num: int, row: Dict[str, Any], file_dir: str, result: ImportResult) -> None:
        """Process a single row and add to result."""
        # Normalize keys to lowercase
        row = {k.lower().strip(): v for k, v in row.items()}
        
        validation_error = self._validate_row(row)
        if validation_error:
            result.errors.append(ImportError(row=row_num, message=validation_error))
            return
        
        name = row.get("name", "").strip()
        animal_type = row.get("animal_type", "").strip().capitalize()
        breed = row.get("breed", "").strip() or None
        age = int(row.get("age", 0))
        health_status = row.get("health_status", "").strip().lower()
        photo_path = row.get("photo", "").strip()
        
        photo_filename = None
        if photo_path:
            photo_result = self._handle_photo(photo_path, file_dir, name)
            if photo_result[0]:  # Success
                photo_filename = photo_result[1]
            else:  # Error
                result.errors.append(ImportError(row=row_num, message=photo_result[1]))
                return
        
        try:
            self.animal_service.add_animal(
                name=name,
                type=animal_type,
                breed=breed,
                age=age,
                health_status=health_status,
                photo=photo_filename,
            )
            result.success_count += 1
        except Exception as e:
            result.errors.append(ImportError(row=row_num, message=f"Database error: {e}"))
    
    def _validate_row(self, row: Dict[str, Any]) -> Optional[str]:
        """Validate a row's data. Returns error message or None."""
        name = row.get("name", "").strip()
        if not name:
            return "Missing required field 'name'"
        
        animal_type = row.get("animal_type", "").strip().lower()
        if not animal_type:
            return "Missing required field 'animal_type'"
        if animal_type not in self.VALID_ANIMAL_TYPES:
            return f"Invalid animal_type '{row.get('animal_type')}'. Must be Dog, Cat, or Other (case-insensitive)"
        
        age_str = str(row.get("age", "")).strip()
        if not age_str:
            return "Missing required field 'age'"
        try:
            age = int(float(age_str))  # Handle "3.0" from Excel
            if age < self.MIN_AGE or age > self.MAX_AGE:
                return f"Invalid age '{age}'. Must be between {self.MIN_AGE} and {self.MAX_AGE}"
        except (ValueError, TypeError):
            return f"Invalid age '{age_str}'. Must be a number from {self.MIN_AGE} to {self.MAX_AGE}"
        
        health_status = row.get("health_status", "").strip().lower()
        if not health_status:
            return "Missing required field 'health_status'"
        if health_status not in self.VALID_HEALTH_STATUSES:
            return f"Invalid health_status '{row.get('health_status')}'. Must be Healthy, Recovering, or Injured (case-insensitive)"
        
        # Breed is optional, no validation needed
        
        return None
    
    def _handle_photo(self, photo_path: str, file_dir: str, animal_name: str) -> Tuple[bool, str]:
        """Handle photo from URL or local path.
        
        Returns:
            Tuple of (success, filename_or_error_message)
        """
        if photo_path.startswith(("http://", "https://")):
            return self._download_photo_from_url(photo_path, animal_name)
        else:
            return self._copy_local_photo(photo_path, file_dir, animal_name)
    
    def _download_photo_from_url(self, url: str, animal_name: str) -> Tuple[bool, str]:
        """Download a photo from a URL and save it.
        
        Returns:
            Tuple of (success, filename_or_error_message)
        """
        try:
            import urllib.request
            import ssl
            
            parsed = urlparse(url)
            url_filename = os.path.basename(parsed.path) or "photo.jpg"
            ext = os.path.splitext(url_filename)[1].lower()
            
            if ext not in app_config.ALLOWED_PHOTO_EXTENSIONS:
                return (False, f"Invalid photo format '{ext}'. Allowed: {', '.join(app_config.ALLOWED_PHOTO_EXTENSIONS)}")
            
            # Download to temp file
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                try:
                    with urllib.request.urlopen(url, timeout=30, context=ctx) as response:
                        shutil.copyfileobj(response, tmp)
                    tmp_path = tmp.name
                except urllib.error.URLError as e:
                    return (False, f"Failed to download photo from URL: {e}")
                except Exception as e:
                    return (False, f"Error downloading photo: {e}")
            
            try:
                # Read the downloaded file
                with open(tmp_path, "rb") as f:
                    image_bytes = f.read()
                
                if len(image_bytes) < 100:
                    return (False, "Downloaded file is too small to be a valid image")
                
                filename = self.file_store.save_bytes(
                    data=image_bytes,
                    original_name=url_filename,
                    custom_name=animal_name,
                )
                return (True, filename)
            finally:
                # Clean up temp file
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                    
        except Exception as e:
            return (False, f"Error processing photo URL: {e}")
    
    def _copy_local_photo(self, photo_path: str, file_dir: str, animal_name: str) -> Tuple[bool, str]:
        """Copy a local photo file to uploads.
        
        Returns:
            Tuple of (success, filename_or_error_message)
        """
        # Resolve relative paths
        if not os.path.isabs(photo_path):
            photo_path = os.path.join(file_dir, photo_path)
        
        # Normalize path
        photo_path = os.path.normpath(photo_path)
        
        if not os.path.isfile(photo_path):
            return (False, f"Photo file not found: {photo_path}")
        
        ext = os.path.splitext(photo_path)[1].lower()
        if ext not in app_config.ALLOWED_PHOTO_EXTENSIONS:
            return (False, f"Invalid photo format '{ext}'. Allowed: {', '.join(app_config.ALLOWED_PHOTO_EXTENSIONS)}")
        
        try:
            # Read the file
            with open(photo_path, "rb") as f:
                image_bytes = f.read()
            
            original_name = os.path.basename(photo_path)
            filename = self.file_store.save_bytes(
                data=image_bytes,
                original_name=original_name,
                custom_name=animal_name,
            )
            return (True, filename)
        except Exception as e:
            return (False, f"Error copying photo: {e}")
    
    @staticmethod
    def get_csv_template_path() -> str:
        """Get the path to the CSV template file."""
        return os.path.join(app_config.ASSETS_DIR, "templates", "animal_import_template.csv")
    
    @staticmethod
    def get_excel_template_path() -> str:
        """Get the path to the Excel template file."""
        return os.path.join(app_config.ASSETS_DIR, "templates", "animal_import_template.xlsx")
    
    @staticmethod
    def generate_csv_template(output_path: str) -> bool:
        """Generate CSV template file with instructions.
        
        Args:
            output_path: Path where the template will be saved
            
        Returns:
            True if successful, False otherwise
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            content = """# ===== PawRes Animal Import Template =====
#
# INSTRUCTIONS:
#
# Column Descriptions:
#   name          - Required. The animal's name
#   animal_type   - Required. Must be: Dog, Cat, or Other (case-insensitive)
#   breed         - Optional. The breed of the animal (e.g., Golden Retriever, Persian)
#   age           - Required. Number from 0 to 21
#                   0 = Under 1 year, 1-20 = Age in years, 21 = Above 20 years
#   health_status - Required. Must be: Healthy, Recovering, or Injured (case-insensitive)
#   photo         - Optional. Web URL or local file path
#
# Supported Photo Formats: .jpg, .jpeg, .png, .gif, .webp
# Lines starting with # are comments and will be ignored.
# Delete the example rows below and add your own data.
#
name,animal_type,breed,age,health_status,photo
Buddy,Dog,Golden Retriever,3,healthy,https://example.com/buddy.jpg
Whiskers,Cat,Persian,0,recovering,
Max,Dog,Labrador,15,injured,
Luna,Cat,,2,healthy,
"""
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(content)
            return True
        except Exception as e:
            print(f"Error generating CSV template: {e}")
            return False
    
    @staticmethod
    def generate_excel_template(output_path: str) -> bool:
        """Generate Excel template file with instructions.
        
        Args:
            output_path: Path where the template will be saved
            
        Returns:
            True if successful, False otherwise
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            wb = Workbook()
            
            # Sheet 1: Animals (Data Sheet)
            ws_data = wb.active
            ws_data.title = "Animals"
            
            headers = ["name", "animal_type", "breed", "age", "health_status", "photo"]
            header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, start=1):
                cell = ws_data.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Example data
            example_data = [
                ("Buddy", "Dog", "Golden Retriever", 3, "healthy", "https://example.com/buddy.jpg"),
                ("Whiskers", "Cat", "Persian", 0, "recovering", "C:\\Photos\\whiskers.png"),
                ("Max", "Dog", "Labrador", 15, "injured", ""),
                ("Luna", "Cat", "", 2, "healthy", "./photos/luna.jpg"),
            ]
            
            for row_num, row_data in enumerate(example_data, start=2):
                for col, value in enumerate(row_data, start=1):
                    ws_data.cell(row=row_num, column=col, value=value)
            
            ws_data.column_dimensions["A"].width = 15
            ws_data.column_dimensions["B"].width = 15
            ws_data.column_dimensions["C"].width = 18
            ws_data.column_dimensions["D"].width = 10
            ws_data.column_dimensions["E"].width = 15
            ws_data.column_dimensions["F"].width = 40
            
            # Sheet 2: Instructions
            ws_inst = wb.create_sheet(title="Instructions")
            instructions = [
                ("PawRes Animal Import Template", True, 16),
                ("", False, 11),
                ("COLUMN DESCRIPTIONS:", True, 12),
                ("", False, 11),
                ("name (Required) - The animal's name", False, 11),
                ("animal_type (Required) - Must be: Dog, Cat, or Other (case-insensitive)", False, 11),
                ("breed (Optional) - The breed of the animal (e.g., Golden Retriever, Persian)", False, 11),
                ("age (Required) - Number 0-21 (0=Under 1 year, 21=Above 20)", False, 11),
                ("health_status (Required) - Must be: Healthy, Recovering, or Injured (case-insensitive)", False, 11),
                ("photo (Optional) - Web URL or local file path", False, 11),
                ("", False, 11),
                ("Supported photo formats: .jpg, .jpeg, .png, .gif, .webp", False, 11),
            ]
            
            for row_num, (text, is_bold, size) in enumerate(instructions, start=1):
                cell = ws_inst.cell(row=row_num, column=1, value=text)
                cell.font = Font(bold=is_bold, size=size)
            
            ws_inst.column_dimensions["A"].width = 70
            
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"Error generating Excel template: {e}")
            return False


__all__ = ["ImportService", "ImportResult", "ImportError"]
